# ========#
#  设置  #
# ========#
# -*- coding:utf-8 -*-
import ctypes
import inspect
import os
import sys
import time
import re
import psutil

from PySide2.QtCore import QObject, Signal
from PySide2.QtGui import QIcon
from openpyxl import load_workbook

from urllib.parse import quote
from PySide2.QtUiTools import QUiLoader
from PySide2.QtWidgets import QApplication, QMessageBox, QTextBrowser, QPlainTextEdit, QPushButton, QLineEdit, \
    QComboBox, QLabel, QProgressBar
import logging
import traceback
from threading import Thread
import Tool
from selenium.webdriver.common.by import By


url = 'https://esi.clarivate.com/'
chrome_path = os.getcwd() + "/Application/v-77/chrome.exe"
root, FILENAME=os.path.split(os.path.abspath(__file__))

url = 'https://esi.clarivate.com/'
url2 = 'http://apps.webofknowledge.com/InboundService.do?product=WOS&Func=Frame&DestFail=http%3A%2F%2Fwww.webofknowledge.com%3FDestParams%3DUT%253DWOS%25253A000359216600008%2526customersID%253DInCites%2526smartRedirect%253Dyes%2526action%253Dretrieve%2526mode%253DFullRecord%2526product%253DCEL%26SrcAuth%3DInCites%26SrcApp%3DTSM_TEST%26DestApp%3DCEL%26e%3DvwtjxiLuDPrhqkxGzMeEeDiukRHn%252FfEx%252FpT5qofj3%252Boj8KOdkGQGQA%253D%253D&SrcApp=TSM_TEST&SrcAuth=InCites&SID=5FqAZgckxxK2sxI7pLR&customersID=InCites&smartRedirect=yes&mode=FullRecord&IsProductCode=Yes&Init=Yes&action=retrieve&UT=WOS%3A000359216600008'
DE_url = ['https://esi.clarivate.com/DocumentsExport.action?exportFile&_dc=1368621151464&groupBy=documentList&start=0&limit=9999&author=&institution=NANJING%20UNIVERSITY%20OF%20INFORMATION%20SCIENCE%20%26%20TECHNOLOGY&territory=&journal=&researchFront=&year=&title=&researchField=&show=highlyCited&sort=%5B%7B%22property%22:%22citations%22,%22direction%22:%22DESC%22%7D%5D&colFilterVal=&exportType=documents&colNames=Accession%20Number,DOI,PMID,Article%20Name,Authors,Source,Research%20Field,Times%20Cited,Countries,Addresses,Institutions,Publication%20Date&fileType=Excel&f=DocumentsExport.xls',
          'https://esi.clarivate.com/DocumentsExport.action?exportFile&_dc=1368621151464&groupBy=documentList&start=0&limit=9999&author=&institution=NANJING%20UNIVERSITY%20OF%20INFORMATION%20SCIENCE%20%26%20TECHNOLOGY&territory=&journal=&researchFront=&year=&title=&researchField=&show=Hot&sort=%5B%7B%22property%22:%22citations%22,%22direction%22:%22DESC%22%7D%5D&colFilterVal=&exportType=documents&colNames=Accession%20Number,DOI,PMID,Article%20Name,Authors,Source,Research%20Field,Times%20Cited,Countries,Addresses,Institutions,Publication%20Date&fileType=Excel&f=DocumentsExport.xls']
#0:HCP,1:HP
TYPE = ['HCP','HP']
Headers={
         "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.120 Safari/537.36",
         "Accept-Language": "zh-CN,zh;q=0.9"
         }
month_short = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
               'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}

def getDEUrl(name, type):
    # type为0: HCP    type为1: HP
    DE_url = [
        'https://esi.clarivate.com/DocumentsExport.action?exportFile&_dc=1368621151464&groupBy=documentList&start=0&limit=9999&author=&institution=NANJING%20UNIVERSITY%20OF%20INFORMATION%20SCIENCE%20%26%20TECHNOLOGY&territory=&journal=&researchFront=&year=&title=&researchField=&show=highlyCited&sort=%5B%7B%22property%22:%22citations%22,%22direction%22:%22DESC%22%7D%5D&colFilterVal=&exportType=documents&colNames=Accession%20Number,DOI,PMID,Article%20Name,Authors,Source,Research%20Field,Times%20Cited,Countries,Addresses,Institutions,Publication%20Date&fileType=Excel&f=DocumentsExport.xls',
        'https://esi.clarivate.com/DocumentsExport.action?exportFile&_dc=1368621151464&groupBy=documentList&start=0&limit=9999&author=&institution=NANJING%20UNIVERSITY%20OF%20INFORMATION%20SCIENCE%20%26%20TECHNOLOGY&territory=&journal=&researchFront=&year=&title=&researchField=&show=Hot&sort=%5B%7B%22property%22:%22citations%22,%22direction%22:%22DESC%22%7D%5D&colFilterVal=&exportType=documents&colNames=Accession%20Number,DOI,PMID,Article%20Name,Authors,Source,Research%20Field,Times%20Cited,Countries,Addresses,Institutions,Publication%20Date&fileType=Excel&f=DocumentsExport.xls']
    if type == 0:
        return 'https://esi.clarivate.com/DocumentsExport.action?exportFile&_dc=1368621151464&groupBy=documentList&start=0&limit=9999&author=&institution={}&territory=&journal=&researchFront=&year=&title=&researchField=&show=highlyCited&sort=%5B%7B%22property%22:%22citations%22,%22direction%22:%22DESC%22%7D%5D&colFilterVal=&exportType=documents&colNames=Accession%20Number,DOI,PMID,Article%20Name,Authors,Source,Research%20Field,Times%20Cited,Countries,Addresses,Institutions,Publication%20Date&fileType=Excel&f=DocumentsExport.xls'.format(quote(name.upper()))
    elif type == 1:
        return 'https://esi.clarivate.com/DocumentsExport.action?exportFile&_dc=1368621151464&groupBy=documentList&start=0&limit=9999&author=&institution={}&territory=&journal=&researchFront=&year=&title=&researchField=&show=Hot&sort=%5B%7B%22property%22:%22citations%22,%22direction%22:%22DESC%22%7D%5D&colFilterVal=&exportType=documents&colNames=Accession%20Number,DOI,PMID,Article%20Name,Authors,Source,Research%20Field,Times%20Cited,Countries,Addresses,Institutions,Publication%20Date&fileType=Excel&f=DocumentsExport.xls'.format(quote(name.upper()))
    else:
        return

def getUrl_new(id):
    SID = '8BpveVXfUipGJ9goxbD'
    url = 'https://www.webofscience.com/wos/woscc/full-record/WOS:{0}?SID={1}'.format(quote(id.upper()), SID)
    return url

def getUrlO(id):
    SID = '5FqAZgckxxK2sxI7pLR'
    url = 'http://apps.webofknowledge.com/InboundService.do?product=WOS&Func=Frame&DestFail=http%3A%2F%2Fwww.webofknowledge.com%3FDestParams%3DUT%253DWOS%25253A000359216600008%2526customersID%253DInCites%2526smartRedirect%253Dyes%2526action%253Dretrieve%2526mode%253DFullRecord%2526product%253DCEL%26SrcAuth%3DInCites%26SrcApp%3DTSM_TEST%26DestApp%3DCEL%26e%3DvwtjxiLuDPrhqkxGzMeEeDiukRHn%252FfEx%252FpT5qofj3%252Boj8KOdkGQGQA%253D%253D&SrcApp=TSM_TEST&SrcAuth=InCites&SID='+SID+'&customersID=InCites&smartRedirect=yes&mode=FullRecord&IsProductCode=Yes&Init=Yes&action=retrieve&UT=WOS%3A'+id
    return url

def getUrl(name):
    url = 'http://gateway.webofknowledge.com/gateway/Gateway.cgi?GWVersion=2&SrcAuth=InCites&SrcApp=TSM_TEST&DestApp=WOS_CPL&DestLinkType=FullRecord&KeyUT='+name
    return url

def getUrl2O(name):
    SID = '5CgcjA1fRe1GUUrmdSV'
    url = 'http://cel.webofknowledge.com/InboundService.do?customersID=InCites&smartRedirect=yes&mode=FullRecord&IsProductCode=Yes&product=CEL&Init=Yes&Func=Frame&action=retrieve&SrcApp=TSM_TEST&SrcAuth=InCites&SID='+SID+'&UT=WOS%3A'+name
    return url

'''
1、还是从老网页下手，找规律（新网页源代码不能获取）
2、新网页，继续钻js 设置document.title window.print()
3、试试插件打印pdf的
'''

# 自定义信号源对象类型，继承自 QObject
class MySignals(QObject):
    # 打印程序运行信息的信号
    log_to_window = Signal(str)

    # 设置按钮状态信号
    set_button = Signal(QPushButton, bool)

    # 设置单行文本信号
    set_lineEdit = Signal(QLineEdit, str)

    # 设置组合框信号
    set_comboBox = Signal(QComboBox)

    # 设置label信号
    set_label = Signal(QLabel, str)

    # 设置进度条信号
    set_progressBar = Signal(QProgressBar, float)
mySignals = MySignals()

# 装饰器，打印错误信息
def errorLog(func):
    def wrapper(self, *args, **kargs):
        try:
            func(self, *args, **kargs)
        except:
            s = traceback.format_exc()
            logging.error(f'\n{s}')
            mySignals.log_to_window.emit(s)
    return wrapper

class MainWindow:
    threadList = []
    info = ''   # 程序运行信息
    temp_fold = root + "\\temp_fold"



    @errorLog
    def __init__(self):
        self.init_log()


        # 将信号的emit（执行方法）绑定为成员方法
        mySignals.log_to_window.connect(self.log_to_window)
        mySignals.set_button.connect(self.set_button)
        mySignals.set_comboBox.connect(self.set_comboBox)
        mySignals.set_lineEdit.connect(self.set_lineEdit)
        mySignals.set_label.connect(self.set_label)
        mySignals.set_progressBar.connect(self.set_progressBar)

        # 加载窗体
        self.ui = QUiLoader().load('ui/main.ui')
        self.ui.pushButton.clicked.connect(self.launch)
        # self.ui.pushButton_2.clicked.connect(self.stop)
        self.ui.comboBox.currentIndexChanged.connect(self.handleSelectionChange)
        self.ui.lineEdit.returnPressed.connect(self.handleReturnPressed)

        self.set_icon()

        # mySignals.set_button.emit(self.ui.pushButton_2, False)


    def set_icon(self):
        appIcon = QIcon("logo.ico")
        self.ui.setWindowIcon(appIcon)

    def init_log(self):
        log_path =  os.getcwd() + '\\log\\'
        log_file = log_path + 'crawler.log'  # 程序日志，记录所有运行信息
        if not os.path.exists(log_path):
            os.makedirs(log_path)
        if not os.path.exists(log_file):
            f = open(log_file, 'w')
            f.close()
        logging.basicConfig(level=logging.DEBUG,  # 控制台打印的日志级别
                            filename=log_file,
                            filemode='a',  # 模式，有w和a，w就是写模式，每次都会重新写日志，覆盖之前的日志
                            # a是追加模式，默认如果不写的话，就是追加模式
                            format='%(asctime)s - %(pathname)s[line:%(lineno)d] - %(levelname)s: %(message)s'  # 日志格式
                            )


    def init_root_chrome(self):
        def run(self):
            if not os.path.exists(self.temp_fold):
                os.mkdir(self.temp_fold)
            # 启动浏览器打开网页,获取最新年月份 （本地的在选择学校后获取）
            self.log_info('启动浏览器...')
            self.chrome = Tool.chromeInit(savePath=self.temp_fold)
            self.log_info('正在加载...')
            while True:
                try:
                    self.chrome.get(url)
                    break
                except:

                    self.log_info("网络错误（esi加载失败）")
                    time.sleep(60)

            update_str = self.chrome.find_element_by_css_selector("#updateDateDatasetESI").text
            self.month = month_short[update_str.split('updated ')[1].split(' ')[0]]
            self.year = int(update_str.split(',')[1].split('.')[0])
            self.log_info('网站当前更新至{}年{}月'.format(self.year, self.month))
            mySignals.set_lineEdit.emit(self.ui.lineEdit_3, '{}年{}月'.format(self.year, self.month))

            # 浏览器到预备检索状态
            self.chrome.find_element(By.CSS_SELECTOR, ".select2-choice").click()  # 点击弹出下拉菜单
            elements = self.chrome.find_elements(By.CSS_SELECTOR, ".select2-result-label")  # 下拉菜单中的元素
            for element in elements:
                if element.text == "Institutions":
                    element.click()  # 选择 Institution
                    break

            self.chrome.find_element(By.CSS_SELECTOR, ".add-filters").click()  # 点击添加过滤器
            self.chrome.find_element(By.CSS_SELECTOR, ".popup-wrapper>ul>li:nth-child(4)").click()  # 点击 Institutions
            self.inputElement = self.chrome.find_element(By.CSS_SELECTOR, ".select2-search-field>input")
            self.log_info('网页加载完毕,请输入学校')

        # 开启线程
        thread = Thread(
            target=run,
            # 注意参数是元组， 如果只有一个参数，后面要有逗号，像这样 args=('参数1',)
            args=(self,),
            daemon=True  # 设置新线程为daemon线程
        )
        self.threadList.append(thread)
        thread.start()

    #################### 与signal绑定的操作界面的方法 ####################

    # @errorLog
    # 由于装饰器中调用了这个方法，若此方法报错，则陷入无限循环报错
    def log_to_window(self, string):
        self.ui.plainTextEdit.appendPlainText(
            '{} {}'.format(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), string))

    def log_info(self, string):
        logging.info(string)
        mySignals.log_to_window.emit(string)

    @errorLog
    def set_button(self, pushButton, flag):
        pushButton.setEnabled(flag)

    @errorLog
    def set_comboBox(self, comboBox):
        # 获取列表 更新comboBox
        comboBox.clear()
        comboBox.addItems(self.result_list)

    @errorLog
    def set_lineEdit(self, lineEdit, string):
        lineEdit.setText(string)

    @errorLog
    def set_label(self, label, string):
        label.setText(string)

    @errorLog
    def set_progressBar(self, progressBar, value):
        progressBar.setValue(value)

    #################### 与signal绑定的操作界面的方法 ####################

    #################### 事件处理 ####################
    @errorLog
    def launch(self):
        mySignals.set_button.emit(self.ui.pushButton, False)
        # mySignals.set_button.emit(self.ui.pushButton_2, True)

        @errorLog
        def run(self, i, ws, process, total):
                temp_fold = './temp_fold_{}'.format(TYPE[i])
                chrome_sub = Tool.chromeInit(temp_fold)

                row_range = ws[7 + process:len(ws['A']) - 2]

                for item in row_range:
                    title = item[3].value
                    rstr = r"[\/\\\:\*\?\"\<\>\|]"  # '/ \ : * ? " < > |'
                    title = re.sub(rstr, '', title)
                    self.log_info('id: {}'.format(item[0].value.split(':')[1]))
                    self.save_pdf(chrome_sub,
                                  getUrlO(item[0].value.split(':')[1]),
                                  '{}/data/{}/{}.{}/{}/'.format(os.getcwd(), self.institution_name, self.year,
                                                                self.month, TYPE[i]) + '\\' + str(process + 1),
                                  '-' + title)
                    if not os.path.exists('./log/{}/{}.{}'.format(self.institution_name, self.year, self.month)):
                        os.makedirs('./log/{}/{}.{}'.format(self.institution_name, self.year, self.month))
                    f = open('./log/{}/{}.{}/{}.log'.format(self.institution_name, self.year, self.month, TYPE[i]), 'w')
                    process += 1
                    f.write(str(process))
                    f.close()

                    if i == 0:  # HCP
                        mySignals.set_progressBar.emit(self.ui.progressBar, process / total * 100)
                    elif i == 1:  # HP
                        mySignals.set_progressBar.emit(self.ui.progressBar_2, process / total * 100)
                    self.log_info('{}:{}进度{}/{}，{:.2f}%\n'.format(self.institution_name, TYPE[i], process, total,
                                                                process / total * 100))
                self.log_info('{}:{}收集完成\n'.format(self.institution_name, TYPE[i]))
                chrome_sub.close()

        for i in range(2):
            try:
                f = open('./log/{}/{}.{}/{}.log'.format(self.institution_name, self.year, self.month, TYPE[i]), 'r')
            except:
                process = 0
                if not os.path.exists('./log/{}/{}.{}/{}'.format(self.institution_name, self.year, self.month, TYPE[i])):
                    os.makedirs('./log/{}/{}.{}/{}'.format(self.institution_name, self.year, self.month, TYPE[i]))
                self.download(self.chrome,
                              self.temp_fold,
                              getDEUrl(self.institution_name, i),
                              './data/{3}/{1}.{2}/{0}/{0}-{1}.{2}.xlsx'.format(TYPE[i], self.year, self.month,self.institution_name)
                              )
                self.log_info('{}:{}总表获取完成'.format(self.institution_name, TYPE[i]))
            else:
                process = int(f.read())
                self.log_info('{}:{}已收集至{}'.format(self.institution_name, TYPE[i], process))
                f.close()
            ws = load_workbook('./data/{3}/{1}.{2}/{0}/{0}-{1}.{2}.xlsx'.format(TYPE[i], self.year, self.month,
                                                                           self.institution_name)).active
            total = ws.max_row - 8
            if i == 0:  # HCP
                mySignals.set_progressBar.emit(self.ui.progressBar, process/total*100)
            elif i == 1:  # HP
                mySignals.set_progressBar.emit(self.ui.progressBar_2, process/total*100)

            if process == total:
                self.log_info('{}:{}已收集完成\n'.format(self.institution_name, TYPE[i]))
            else:
                # 开启线程
                thread = Thread(
                    target=run,
                    # 注意参数是元组， 如果只有一个参数，后面要有逗号，像这样 args=('参数1',)
                    args=(self, i, ws, process, total),
                    daemon=True  # 设置新线程为daemon线程
                )
                self.threadList.append(thread)
                thread.start()


    # @errorLog
    # def stop(self):
    #     mySignals.set_button.emit(self.ui.pushButton, True)
    #     mySignals.set_button.emit(self.ui.pushButton_2, False)
    #     for thread in self.threadList:
    #         stop_thread(thread)


    @errorLog
    def handleSelectionChange(self, index):
        @errorLog
        def run(self):
            # 根据选中的，设置单行文本的文字
            self.institution_name = self.ui.comboBox.currentText()
            mySignals.set_lineEdit.emit(self.ui.lineEdit, self.institution_name)

            # 获取本地最新更新时间
            if os.path.exists('./log/{}'.format(self.institution_name)):
                local_latest_str = Tool.sort_fold('./log/{}'.format(self.institution_name))
                if local_latest_str == '文件夹为空':
                    local_latest_str = '暂无该学校记录'
                else:
                    local_latest_str = '{}年{}月'.format(local_latest_str.split('.')[0], local_latest_str.split('.')[1])
            else:
                os.makedirs('./log/{}'.format(self.institution_name))
                local_latest_str = '暂无该学校记录'
            mySignals.set_lineEdit.emit(self.ui.lineEdit_2, local_latest_str)

            # 获取收集进度
            for i in range(2):
                try:
                    f = open('./log/{}/{}.{}/{}.log'.format(self.institution_name, self.year, self.month, TYPE[i]), 'r')
                except:
                    total = 1
                    process = 0
                else:
                    process = int(f.read())
                    self.log_info('{}:{}已收集至{}'.format(self.institution_name, TYPE[i], process))
                    f.close()
                    ws = load_workbook('./data/{3}/{1}.{2}/{0}/{0}-{1}.{2}.xlsx'.format(TYPE[i], self.year, self.month, self.institution_name)).active
                    total = ws.max_row - 8
                    if process == total:
                        self.log_info('{}:{}已收集完成'.format(self.institution_name, TYPE[i]))
                        continue

                if i == 0:  # HCP
                    mySignals.set_progressBar.emit(self.ui.progressBar, process / total)
                elif i == 1:  # HP
                    mySignals.set_progressBar.emit(self.ui.progressBar_2, process / total)


        # 开启线程
        thread = Thread(
            target=run,
            args=(self,),
            daemon=True  # 设置新线程为daemon线程
        )
        self.threadList.append(thread)
        thread.start()


    @errorLog
    def handleReturnPressed(self):
        @errorLog
        def run(self):
            # 获取文字，发起检索
            last_key = self.inputElement.get_attribute('value') # 上一次搜索信息
            self.log_info('界面输入框:{}'.format(last_key))
            if last_key != self.ui.lineEdit.text():
                self.log_info('发起搜索...')
                #关键词和上一次不一样才发起搜索
                self.inputElement.clear()  # 清除输入框已有的字符串
                self.inputElement.send_keys(self.ui.lineEdit.text())
                time.sleep(1)
                self.result_label_list = self.chrome.find_elements(By.CSS_SELECTOR, ".select2-results-dept-0.select2-result.select2-result-selectable>div")

            self.result_list = []
            if len(self.result_label_list) != 0:
                for element in self.result_label_list:
                    self.result_list.insert(0, element.text)
            mySignals.set_comboBox.emit(self.ui.comboBox)

        # 开启线程
        thread = Thread(
            target=run,
            args=(self,), # 只有一个参数的时候要加,使其变成元组
            daemon=True  # 设置新线程为daemon线程
        )
        self.threadList.append(thread)
        thread.start()

    def close_window(self):
        self.chrome.close()

    #################### 事件处理 ####################

    #################### 下载方法 ####################

    def download(self, chrome, temp_fold, urls, dst_file):
        while True:
            try:
                self.log_info('访问此链接下载:{}'.format(urls))
                Tool.clean_fold(temp_fold)
                chrome.get(urls)
            except Exception:
                self.log_info("网络错误（下载失败）\n")
                time.sleep(60)
                continue

            try:
                # 等待下载完成 确保中间文件(.tmp .crdownload)完全转好
                while (len(os.listdir(temp_fold)) == 0
                       or os.listdir(temp_fold)[0].split('.')[-1] == 'tmp'
                       or os.listdir(temp_fold)[0].split('.')[-1] == 'crdownload'): time.sleep(0.1)

                Tool.movefile(temp_fold + '\\' + Tool.sort_fold(temp_fold), dst_file)
                time.sleep(0.1)  # 稍微控制时间，防止反爬
                break
            except:
                self.log_info('文件移动转换错误... 重新下载...\n')
    
    # 用chrome.exe将本地html转pdf
    # src_file, dst_file 须为绝对路径
    def html2pdf(self, chrome_path, src_file, dst_file):
        chrome_path_filted = chrome_path.replace('(', '^(') \
            .replace(')', '^)')

        # 文件名里会有空格(cmd不能识别)，需要""
        src_file = r'"' + src_file + r'"'
        dst_file = r'"' + dst_file + r'"'

        cmd_str = chrome_path_filted + ' ' + '--headless --disable-gpu --print-to-pdf=' + dst_file + ' ' + src_file
        cmd_str = cmd_str.replace('/', '\\')
        # os.system(cmd_str)
        psutil.Popen(cmd_str, shell=True)

        self.log_info(cmd_str)

    first_enter = True
    def save_pdf(self, chrome, urls, dst_path, title, chrome_path=chrome_path):
        while True:
            try:
                self.log_info('访问此链接下载html:{}'.format(urls))
                chrome.get(urls)

                # 如果是第一次打开 web of science 链接 网页有些不必要的提示信息
                if self.first_enter == True:
                    time.sleep(0.5)
                    chrome.get(urls)
                    first_enter = False

                # chrome.execute_script('document.title = "test"; window.print()')

                text = chrome.page_source
                break
            except Exception:
                self.log_info("网络错误（下载失败）\n")
                time.sleep(60)
                continue
        title = title[:30]  # 防止文件名过长，就30够了
        with open(dst_path + title + '.html', 'w', encoding='utf-8') as f:
            self.log_info(dst_path + title + '.html')
            f.write('<head><meta charset="UTF-8"></head>' + text)
        time.sleep(0.5)  # 防反爬

        self.html2pdf(chrome_path, dst_path + title + '.html', dst_path + title + '.pdf')

        #################### 下载方法 ####################



if __name__ == "__main__":
    try:
    # 确保能获取到所有报错信息

        # 启动窗体
        app = QApplication([])
        main_window = MainWindow()
        main_window.ui.show()
        main_window.init_root_chrome()
        app.exec_()
        app.aboutToQuit.connect(main_window.close_window)

        main_window.chrome.close()

    except:
        s = traceback.format_exc()
        logging.error(f'\n{s}')